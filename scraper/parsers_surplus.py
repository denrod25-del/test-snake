"""
parsers_surplus.py
------------------
HTML parsers for county Clerk surplus-funds pages.

Each parser takes raw HTML and the source URL, and returns a list of dicts
with this shape:

    {
        "sale_date":      "YYYY-MM-DD" or None,
        "parcel_id":      "..." or None,
        "property_addr":  "..." or None,
        "prior_owner":    "..." or None,
        "surplus_amount": float or None,
        "status":         "unclaimed" | "claim_filed" | "paid" | "escheated",
        "claim_deadline": "YYYY-MM-DD" or None,
        "source_url":     str
    }

Surplus pages vary wildly across the 67 counties. The parsers here cover the
most common platforms; counties with PDF-only or unique formats fall back to
`manual_surplus.json`.
"""

from __future__ import annotations

import re
from datetime import datetime, timedelta
from typing import List, Dict, Any, Optional

from bs4 import BeautifulSoup
from dateutil import parser as dateparser


class NoDataError(Exception):
    """Raised when a parser ran but found no records on the page."""


# --------------------------------------------------------------------------
# helpers
# --------------------------------------------------------------------------
_MONEY_RE = re.compile(r"\$?\s*([\d,]+(?:\.\d{1,2})?)")
_DATE_RE = re.compile(r"\b(\d{1,2}[/-]\d{1,2}[/-]\d{2,4}|\w+\s+\d{1,2},?\s+\d{4})\b")
_FEE_NOISE = re.compile(
    r"application fee|sheriff fee|postage|photocopy|advertising fee|"
    r"holiday schedule|contact us|opening\b",
    re.I,
)

_STATUS_KEYWORDS = {
    "claim_filed": ("claim filed", "claim received", "pending"),
    "paid": ("paid", "disbursed", "released"),
    "escheated": ("escheated", "forfeited"),
}

# Reject fee-schedule tables that mention tiny statutory amounts.
_MIN_SURPLUS_AMOUNT = 25.0


def _parse_money(text: str) -> Optional[float]:
    if not text:
        return None
    # Prefer explicit $ amounts, then decimal money-looking numbers.
    # Do NOT match bare integers (e.g. date fragments like 08/27/25 → 08).
    for rx in (
        re.compile(r"\$\s*([\d,]+(?:\.\d{1,2})?)"),
        re.compile(r"(?<![/\d])([\d,]+\.\d{2})(?!\d)"),
    ):
        m = rx.search(text)
        if not m:
            continue
        try:
            return float(m.group(1).replace(",", ""))
        except ValueError:
            continue
    return None


def _parse_date(text: str) -> Optional[str]:
    if not text:
        return None
    m = _DATE_RE.search(text)
    if not m:
        return None
    try:
        return dateparser.parse(m.group(1)).date().isoformat()
    except (ValueError, OverflowError):
        return None


def _detect_status(row_text: str) -> str:
    txt = row_text.lower()
    for status, keywords in _STATUS_KEYWORDS.items():
        if any(kw in txt for kw in keywords):
            return status
    return "unclaimed"


def _claim_deadline_from_sale(sale_iso: Optional[str]) -> Optional[str]:
    """Per FS 197.582, surplus claims must be made within 120 days of sale."""
    if not sale_iso:
        return None
    try:
        sale = datetime.strptime(sale_iso, "%Y-%m-%d").date()
        return (sale + timedelta(days=120)).isoformat()
    except ValueError:
        return None


def _header_looks_like_surplus(headers: List[str]) -> bool:
    joined = " ".join(headers)
    return any(
        token in joined
        for token in ("surplus", "unclaimed", "excess", "overbid", "proceeds")
    )


# --------------------------------------------------------------------------
# parsers
# --------------------------------------------------------------------------
def parse_clerk_html_table(html: str, source_url: str) -> List[Dict[str, Any]]:
    """
    Generic parser for Clerk pages that publish surplus as an HTML table.
    Prefers tables whose headers mention surplus / unclaimed / excess funds.
    """
    soup = BeautifulSoup(html, "lxml")
    tables = soup.find_all("table")
    if not tables:
        raise NoDataError("no <table> elements found")

    candidates = []
    for table in tables:
        rows = table.find_all("tr")
        if len(rows) < 2:
            continue
        headers = [_normalize(th.get_text()) for th in rows[0].find_all(["th", "td"])]
        if not headers:
            continue
        score = len(rows)
        if _header_looks_like_surplus(headers):
            score += 1000
        candidates.append((score, headers, rows))

    if not candidates:
        raise NoDataError("table has no data rows")

    candidates.sort(key=lambda x: x[0], reverse=True)
    _score, headers, rows = candidates[0]
    if not _header_looks_like_surplus(headers):
        raise NoDataError("largest table is not a surplus list (missing surplus headers)")

    col_map = {
        "sale_date": _find_col(headers, "sale date", "sale", "date"),
        "parcel_id": _find_col(headers, "parcel", "folio", "tax id", "case", "file"),
        "property_addr": _find_col(headers, "property", "address", "situs", "location"),
        "prior_owner": _find_col(headers, "owner", "name", "prior", "defendant"),
        "surplus_amount": _find_col(headers, "surplus", "amount", "balance", "funds", "excess"),
        "status": _find_col(headers, "status", "claim"),
        "claim_deadline": _find_col(headers, "deadline", "1 year", "expire", "claim by"),
    }

    out: List[Dict[str, Any]] = []
    for tr in rows[1:]:
        cells = [td.get_text(" ", strip=True) for td in tr.find_all(["td", "th"])]
        # Drop empty spacer cells some CMS tables insert
        cells_compact = [c for c in cells if c]
        if len(cells_compact) < 2:
            continue
        row_text = " ".join(cells_compact)
        if _FEE_NOISE.search(row_text) and "surplus" not in row_text.lower():
            continue

        sale_iso = _parse_date(_get(cells, col_map["sale_date"]) or _get(cells_compact, col_map["sale_date"]) or row_text)
        amount = _parse_money(
            _get(cells, col_map["surplus_amount"])
            or _get(cells_compact, col_map["surplus_amount"])
            or row_text
        )

        if amount is None or amount < _MIN_SURPLUS_AMOUNT:
            continue
        if not sale_iso:
            continue

        deadline = _parse_date(_get(cells, col_map["claim_deadline"]) or "")
        if not deadline:
            deadline = _claim_deadline_from_sale(sale_iso)

        parcel = _clean(
            _get(cells, col_map["parcel_id"]) or _get(cells_compact, col_map["parcel_id"])
        )
        owner = _clean(
            _get(cells, col_map["prior_owner"]) or _get(cells_compact, col_map["prior_owner"])
        )
        addr = _clean(
            _get(cells, col_map["property_addr"]) or _get(cells_compact, col_map["property_addr"])
        )
        # Avoid treating owner name as address when columns collapse
        if addr and owner and addr == owner:
            addr = None

        out.append({
            "sale_date": sale_iso,
            "parcel_id": parcel,
            "property_addr": addr,
            "prior_owner": owner,
            "surplus_amount": amount,
            "status": _detect_status(_get(cells, col_map["status"]) or row_text),
            "claim_deadline": deadline,
            "source_url": source_url,
        })

    if not out:
        raise NoDataError("table parsed but no usable surplus rows")
    return out


def parse_realauction_surplus(html: str, source_url: str) -> List[Dict[str, Any]]:
    """
    Realauction (realtaxdeed.com) renders surplus through a paginated table
    or a 'Surplus Funds' tab. The structure is reasonably consistent across
    counties — look for rows whose cells contain a sale date and a $ amount.
    """
    soup = BeautifulSoup(html, "lxml")
    out: List[Dict[str, Any]] = []
    for tr in soup.find_all("tr"):
        cells = [td.get_text(" ", strip=True) for td in tr.find_all("td")]
        if len(cells) < 3:
            continue
        line = " | ".join(cells)
        if "$" not in line:
            continue
        if _FEE_NOISE.search(line):
            continue
        sale_iso = _parse_date(line)
        amount = _parse_money(line)
        if amount is None or amount < _MIN_SURPLUS_AMOUNT or not sale_iso:
            continue
        parcel = next((c for c in cells if re.search(r"\d{2,}-\d{2,}", c)), None)
        out.append({
            "sale_date": sale_iso,
            "parcel_id": parcel,
            "property_addr": _longest(cells),
            "prior_owner": None,
            "surplus_amount": amount,
            "status": _detect_status(line),
            "claim_deadline": _claim_deadline_from_sale(sale_iso),
            "source_url": source_url,
        })
    if not out:
        raise NoDataError("no surplus rows found on realauction page")
    return out


def parse_gulf_surplus_cards(html: str, source_url: str) -> List[Dict[str, Any]]:
    """
    Gulf County publishes surplus cards as labeled plain text (Sale Date / Case /
    Parcel / Owner / Location / $amount), not an HTML table.
    """
    text = BeautifulSoup(html, "lxml").get_text("\n", strip=True)
    parts = re.split(r"(?=Sale Date\b)", text)
    out: List[Dict[str, Any]] = []
    for part in parts:
        if "Parcel ID" not in part or "$" not in part:
            continue
        if not re.search(r"\bsurplus\b", part, re.I):
            continue
        sale_iso = _parse_date(_field_after(part, r"Sale Date"))
        parcel = _clean(_field_after(part, r"Parcel ID"))
        case_no = _clean(_field_after(part, r"Case No\.?"))
        owner = _clean(_field_after(part, r"Owner"))
        location = _clean(_field_after(part, r"Location"))
        amount = _parse_money(part)
        if not sale_iso or amount is None or amount < _MIN_SURPLUS_AMOUNT:
            continue
        # Location may span 2 lines (street + city); keep first line-ish blob
        if location and "Applicant" in location:
            location = None
        out.append({
            "sale_date": sale_iso,
            "parcel_id": parcel or case_no,
            "property_addr": location,
            "prior_owner": owner,
            "surplus_amount": amount,
            "status": "unclaimed",
            "claim_deadline": _claim_deadline_from_sale(sale_iso),
            "source_url": source_url,
        })
    if not out:
        raise NoDataError("no Gulf surplus cards found")
    return out


def parse_clerk_pdf(pdf_bytes: bytes, source_url: str) -> List[Dict[str, Any]]:
    """
    Extract surplus rows from Clerk PDF reports (text-based PDFs).
    Dispatches by report title / column layout.
    """
    try:
        import fitz  # PyMuPDF
    except ImportError as e:
        raise NoDataError("pymupdf is required for clerk_pdf parser") from e

    if not pdf_bytes or not pdf_bytes.startswith(b"%PDF"):
        raise NoDataError("payload is not a PDF")

    doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    pages = [doc[i].get_text("text") for i in range(doc.page_count)]
    text = "\n".join(pages)
    lower = text.lower()

    if "tax deeds surplus funds report" in lower or "sale number" in lower and "current balance" in lower:
        rows = _parse_marion_surplus_pdf(text, source_url)
    elif "tax deeds surplus funds available" in lower and "property id" in lower:
        rows = _parse_osceola_surplus_pdf(text, source_url)
    elif "excess proceeds" in lower or ("unclaimed" in lower and "property id" in lower):
        rows = _parse_collier_surplus_pdf(text, source_url)
    elif ("file#" in lower or "file #" in lower) and "surplus" in lower and "sale date" in lower:
        rows = _parse_santa_rosa_surplus_pdf(text, source_url)
    else:
        rows = _parse_generic_surplus_pdf(text, source_url)

    if not rows:
        raise NoDataError("PDF parsed but no usable surplus rows")
    return rows


_MARION_ROW_RE = re.compile(
    r"(?P<sale_num>\d{5,6})\s+"
    r"(?P<sale_date>\d{4}-\d{2}-\d{2})\s+"
    r"(?P<tax>\S+)\s+"
    r"(?P<parcel>\S+)\s+"
    r"\$?\s*(?P<amount>[\d,]+\.\d{2})",
    re.M,
)

_COLLIER_ROW_RE = re.compile(
    r"(?P<sale_date>\d{1,2}/\d{1,2}/\d{4})\s+"
    r"(?P<tda>\d{4,6})\s+"
    r"\$?\s*(?P<amount>[\d,]+\.\d{2})\s+"
    r"(?P<deadline>\d{1,2}/\d{1,2}/\d{4})\s+"
    r"(?P<rest>.*?)(?=\d{1,2}/\d{1,2}/\d{4}\s+\d{4,6}\s+\$?[\d,]+\.\d{2}|\Z)",
    re.S,
)


def _parse_marion_surplus_pdf(text: str, source_url: str) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []
    for m in _MARION_ROW_RE.finditer(text):
        amount = _parse_money(m.group("amount"))
        sale_iso = m.group("sale_date")
        if amount is None or amount < _MIN_SURPLUS_AMOUNT:
            continue
        out.append({
            "sale_date": sale_iso,
            "parcel_id": _clean(m.group("parcel")),
            "property_addr": None,
            "prior_owner": None,
            "surplus_amount": amount,
            "status": "unclaimed",
            "claim_deadline": _claim_deadline_from_sale(sale_iso),
            "source_url": source_url,
        })
    return out


# Osceola: sale date is printed once, then multiple Tax Deed # / Cert # / $ / Property ID rows.


def _parse_osceola_surplus_pdf(text: str, source_url: str) -> List[Dict[str, Any]]:
    """Osceola Clerk 'Tax Deeds Surplus Funds Available' PDF (fields often one-per-line)."""
    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
    out: List[Dict[str, Any]] = []
    current_sale: Optional[str] = None
    i = 0
    while i < len(lines):
        line = lines[i]
        if re.fullmatch(r"\d{1,2}/\d{1,2}/\d{4}", line):
            current_sale = _parse_date(line)
            i += 1
            continue
        # Expect: Tax Deed # / Cert # / $amount / Property ID #
        if (
            current_sale
            and re.fullmatch(r"\d{1,4}-\d{4}", line)
            and i + 3 < len(lines)
            and re.fullmatch(r"\d+", lines[i + 1])
            and _parse_money(lines[i + 2]) is not None
        ):
            amount = _parse_money(lines[i + 2])
            parcel = _clean(lines[i + 3])
            if amount is not None and amount >= _MIN_SURPLUS_AMOUNT and parcel:
                out.append({
                    "sale_date": current_sale,
                    "parcel_id": parcel,
                    "property_addr": None,
                    "prior_owner": None,
                    "surplus_amount": amount,
                    "status": "unclaimed",
                    "claim_deadline": _claim_deadline_from_sale(current_sale),
                    "source_url": source_url,
                })
            i += 4
            continue
        i += 1
    return out


def _parse_santa_rosa_surplus_pdf(text: str, source_url: str) -> List[Dict[str, Any]]:
    """Santa Rosa Clerk tax-deed surplus PDF (one field per line)."""
    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
    # Drop header labels
    while lines and lines[0].upper() in {
        "FILE#", "FILE #", "SURPLUS", "SALE DATE", "PAYEE", "ADDRESS", "CITY", "ST", "ZIP"
    }:
        lines.pop(0)

    out: List[Dict[str, Any]] = []
    i = 0
    while i < len(lines):
        if not re.fullmatch(r"\d{6,8}", lines[i]):
            i += 1
            continue
        file_id = lines[i]
        i += 1
        if i >= len(lines):
            break
        amount = _parse_money(lines[i])
        if amount is None:
            continue
        i += 1
        if i < len(lines) and lines[i] == "$":
            i += 1
        if i >= len(lines):
            break
        sale_iso = _parse_date(lines[i])
        if not sale_iso:
            continue
        i += 1
        owner = _clean(lines[i]) if i < len(lines) else None
        i += 1
        if i < len(lines) and lines[i].upper().startswith("C/O"):
            i += 1  # skip care-of
        addr = None
        if i < len(lines) and not re.fullmatch(r"\d{6,8}", lines[i]):
            addr = _clean(lines[i])
            i += 1
        # Skip city / state / zip until next file id
        while i < len(lines) and not re.fullmatch(r"\d{6,8}", lines[i]):
            i += 1
        if amount < _MIN_SURPLUS_AMOUNT:
            continue
        out.append({
            "sale_date": sale_iso,
            "parcel_id": file_id,
            "property_addr": addr,
            "prior_owner": owner,
            "surplus_amount": amount,
            "status": "unclaimed",
            "claim_deadline": _claim_deadline_from_sale(sale_iso),
            "source_url": source_url,
        })
    return out


def _parse_collier_surplus_pdf(text: str, source_url: str) -> List[Dict[str, Any]]:
    compact = re.sub(r"[ \t]+", " ", text)
    out: List[Dict[str, Any]] = []
    for m in _COLLIER_ROW_RE.finditer(compact):
        amount = _parse_money(m.group("amount"))
        sale_iso = _parse_date(m.group("sale_date"))
        if amount is None or amount < _MIN_SURPLUS_AMOUNT or not sale_iso:
            continue
        rest = m.group("rest") or ""
        status = "claim_filed" if re.search(r"1st claim|claim pending|claims filed", rest, re.I) else "unclaimed"
        prop_ids = re.findall(r"\b\d{8,}\b", rest)
        parcel = prop_ids[-1] if prop_ids else m.group("tda")
        owner = None
        for line in rest.splitlines():
            line = line.strip()
            if not line or re.search(r"claim|deadline|address|city|state|legal|property id", line, re.I):
                continue
            if re.match(r"^\d", line):
                continue
            owner = _clean(line)
            break
        deadline = _parse_date(m.group("deadline")) or _claim_deadline_from_sale(sale_iso)
        out.append({
            "sale_date": sale_iso,
            "parcel_id": parcel,
            "property_addr": None,
            "prior_owner": owner,
            "surplus_amount": amount,
            "status": status,
            "claim_deadline": deadline,
            "source_url": source_url,
        })
    return out


def _parse_generic_surplus_pdf(text: str, source_url: str) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []
    for line in text.splitlines():
        if "$" not in line and not re.search(r"\d+\.\d{2}", line):
            continue
        if _FEE_NOISE.search(line):
            continue
        sale_iso = _parse_date(line)
        amount = _parse_money(line)
        if not sale_iso or amount is None or amount < _MIN_SURPLUS_AMOUNT:
            continue
        parcel = None
        m = re.search(r"\b(\d{2,}[\d\-./]{4,})\b", line)
        if m:
            parcel = m.group(1)
        out.append({
            "sale_date": sale_iso,
            "parcel_id": parcel,
            "property_addr": None,
            "prior_owner": None,
            "surplus_amount": amount,
            "status": _detect_status(line),
            "claim_deadline": _claim_deadline_from_sale(sale_iso),
            "source_url": source_url,
        })
    return out


def _field_after(block: str, label: str) -> Optional[str]:
    m = re.search(rf"{label}\s*\n+(.+)", block, re.I)
    if not m:
        return None
    return m.group(1).strip()


def discover_pdf_url(html: str, page_url: str, pattern: str) -> Optional[str]:
    """Pick the newest matching PDF/XLSX href from an index/listing page."""
    from urllib.parse import urljoin

    rx = re.compile(pattern, re.I)
    soup = BeautifulSoup(html, "lxml")
    matches = []
    for a in soup.find_all("a", href=True):
        href = a["href"]
        label = a.get_text(" ", strip=True) or ""
        if rx.search(href) or rx.search(label):
            matches.append(urljoin(page_url, href))
    if not matches:
        return None
    matches = sorted(set(matches))
    return matches[-1]


def parse_clerk_csv(csv_bytes: bytes, source_url: str) -> List[Dict[str, Any]]:
    """
    Parse Clerk surplus CSV / Google Sheet exports (Sumter-style).
    Expects headers like PROPERTY OWNER, SALE DATE, AMOUNT OF SURPLUS, PARCEL #.
    Address may appear on the following blank-keyed row.
    """
    import csv
    import io

    if not csv_bytes:
        raise NoDataError("empty CSV payload")

    text = csv_bytes.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if not rows:
        raise NoDataError("CSV has no rows")

    header_idx = None
    headers: List[str] = []
    for i, row in enumerate(rows):
        joined = " ".join(row).lower()
        if "sale date" in joined and ("surplus" in joined or "parcel" in joined):
            header_idx = i
            headers = [_normalize(c) for c in row]
            break
    if header_idx is None:
        raise NoDataError("CSV missing surplus header row")

    col = {
        "owner": _find_col(headers, "property owner", "owner", "name"),
        "sale_date": _find_col(headers, "sale date", "sale"),
        "amount": _find_col(headers, "amount of surplus", "surplus", "amount"),
        "parcel": _find_col(headers, "parcel", "folio"),
        "claims": _find_col(headers, "claim"),
    }

    out: List[Dict[str, Any]] = []
    i = header_idx + 1
    while i < len(rows):
        cells = rows[i]
        sale_iso = _parse_date(_get(cells, col["sale_date"]) or "")
        amount = _parse_money(str(_get(cells, col["amount"]) or ""))
        parcel = _clean(str(_get(cells, col["parcel"]) or ""))
        owner = _clean(str(_get(cells, col["owner"]) or ""))
        claims = str(_get(cells, col["claims"]) or "")
        # Address continuation on next row (owner cell only)
        addr = None
        if i + 1 < len(rows):
            nxt = rows[i + 1]
            nxt_sale = _parse_date(_get(nxt, col["sale_date"]) or "")
            nxt_amt = _parse_money(str(_get(nxt, col["amount"]) or ""))
            if not nxt_sale and not nxt_amt:
                maybe_addr = _clean(str(_get(nxt, col["owner"]) or nxt[0] if nxt else ""))
                if maybe_addr and not maybe_addr.lower().startswith("list last"):
                    addr = maybe_addr
                    i += 1
        if sale_iso and amount is not None and amount >= _MIN_SURPLUS_AMOUNT:
            status = "claim_filed" if re.search(r"claim", claims, re.I) and "no claim" not in claims.lower() else "unclaimed"
            if re.search(r"no claim", claims, re.I):
                status = "unclaimed"
            out.append({
                "sale_date": sale_iso,
                "parcel_id": parcel,
                "property_addr": addr,
                "prior_owner": owner,
                "surplus_amount": amount,
                "status": status,
                "claim_deadline": _claim_deadline_from_sale(sale_iso),
                "source_url": source_url,
            })
        i += 1

    if not out:
        raise NoDataError("CSV parsed but no usable surplus rows")
    return out


def parse_clerk_xlsx(xlsx_bytes: bytes, source_url: str) -> List[Dict[str, Any]]:
    """
    Parse Clerk surplus workbooks (Pasco-style monthly sheets with
    DATE RECEIVED / TDA # / ORIGINAL OWNER / PARCEL ID # / ACTUAL BALANCE / BALANCE).
    """
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise NoDataError("openpyxl is required for clerk_xlsx parser") from e

    import io

    if not xlsx_bytes or xlsx_bytes[:2] != b"PK":
        raise NoDataError("payload is not an XLSX workbook")

    wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    # sheet order is chronological in Pasco file; later sheets win on dedupe
    by_key: Dict[str, Dict[str, Any]] = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        header_idx = None
        headers: List[str] = []
        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            vals = [("" if v is None else str(v)).strip() for v in row]
            joined = " ".join(vals).lower()
            if "parcel" in joined and ("balance" in joined or "owner" in joined or "tda" in joined):
                header_idx = i
                headers = [_normalize(v) for v in vals]
                break
        if header_idx is None:
            continue

        col = {
            "sale_date": _find_col(headers, "date received", "sale date", "sale", "date"),
            "parcel_id": _find_col(headers, "parcel id", "parcel", "folio"),
            "case_id": _find_col(headers, "tda", "case", "file"),
            "prior_owner": _find_col(headers, "original owner", "owner", "name"),
            "actual": _find_col(headers, "actual balance", "surplus", "excess"),
            "balance": _find_col(headers, "balance"),
            "date_paid": _find_col(headers, "date paid", "paid date"),
            "amount_paid": _find_col(headers, "amount paid"),
        }

        for row in ws.iter_rows(min_row=header_idx + 1, values_only=True):
            cells = list(row)
            parcel = _clean(str(_get(cells, col["parcel_id"]) or ""))
            case_id = _clean(str(_get(cells, col["case_id"]) or ""))
            owner = _clean(str(_get(cells, col["prior_owner"]) or ""))

            bal_raw = _get(cells, col["balance"])
            act_raw = _get(cells, col["actual"])
            amount = _excel_money(bal_raw)
            if amount is None:
                amount = _excel_money(act_raw)
            if amount is None or amount < _MIN_SURPLUS_AMOUNT:
                continue

            sale_raw = _get(cells, col["sale_date"])
            sale_iso = _excel_date(sale_raw)
            if not sale_iso:
                continue

            paid_date = _excel_date(_get(cells, col["date_paid"]))
            status = "paid" if amount < 1 and paid_date else "unclaimed"

            key = f"{case_id or ''}|{parcel or ''}|{sale_iso}"
            by_key[key] = {
                "sale_date": sale_iso,
                "parcel_id": parcel or case_id,
                "property_addr": None,
                "prior_owner": owner,
                "surplus_amount": amount,
                "status": status,
                "claim_deadline": _claim_deadline_from_sale(sale_iso),
                "source_url": source_url,
            }

    out = list(by_key.values())
    if not out:
        raise NoDataError("XLSX parsed but no usable surplus rows")
    return out


def _excel_money(val: Any) -> Optional[float]:
    if val is None or val == "":
        return None
    if isinstance(val, (int, float)):
        return float(val)
    return _parse_money(str(val))


def _excel_date(val: Any) -> Optional[str]:
    if val is None or val == "":
        return None
    if hasattr(val, "date"):
        try:
            return val.date().isoformat()
        except Exception:
            pass
    return _parse_date(str(val))


# --------------------------------------------------------------------------
# small helpers
# --------------------------------------------------------------------------
def _normalize(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").strip()).lower()


def _find_col(headers: List[str], *needles: str) -> Optional[int]:
    # Prefer longer / more specific needles first via exact-ish match order
    for n in needles:
        for i, h in enumerate(headers):
            if n in h:
                return i
    return None


def _get(cells: List[Any], i: Optional[int]) -> Optional[Any]:
    if i is None or i >= len(cells):
        return None
    return cells[i]


def _clean(s: Optional[str]) -> Optional[str]:
    if not s:
        return None
    return re.sub(r"\s+", " ", s).strip() or None


def _longest(cells: List[str]) -> Optional[str]:
    if not cells:
        return None
    return max(cells, key=len)


def parse_calhoun_overbids(html: str, source_url: str) -> List[Dict[str, Any]]:
    """
    Calhoun Clerk embeds surplus as a Vue prop:
      <tax-deed-overbids :taxdeeds="[{...}, ...]">
    """
    import json

    m = re.search(r':taxdeeds="(\[.*?\])"', html, re.S)
    if not m:
        raise NoDataError("no tax-deed-overbids JSON found")
    raw = (
        m.group(1)
        .replace("&quot;", '"')
        .replace("&#34;", '"')
        .replace("&amp;", "&")
        .replace("&#039;", "'")
    )
    try:
        items = json.loads(raw)
    except json.JSONDecodeError as e:
        raise NoDataError(f"Calhoun overbids JSON invalid: {e}") from e

    out: List[Dict[str, Any]] = []
    for item in items or []:
        amount = _parse_money(str(item.get("balance") or ""))
        sale_iso = _parse_date(str(item.get("sale_date") or ""))
        if amount is None or amount < _MIN_SURPLUS_AMOUNT or not sale_iso:
            continue
        deadline = _parse_date(str(item.get("unclaimed_date") or "")) or _claim_deadline_from_sale(sale_iso)
        parcel = _clean(str(item.get("parcel") or item.get("file") or ""))
        owner = _clean(str(item.get("owner") or ""))
        out.append({
            "sale_date": sale_iso,
            "parcel_id": parcel,
            "property_addr": None,
            "prior_owner": owner,
            "surplus_amount": amount,
            "status": "unclaimed",
            "claim_deadline": deadline,
            "source_url": source_url,
        })
    if not out:
        raise NoDataError("Calhoun overbids parsed but no usable rows")
    return out


PARSERS = {
    "clerk_html_table": parse_clerk_html_table,
    "realauction_surplus": parse_realauction_surplus,
    "gulf_surplus_cards": parse_gulf_surplus_cards,
    "clerk_pdf": parse_clerk_pdf,
    "clerk_xlsx": parse_clerk_xlsx,
    "clerk_csv": parse_clerk_csv,
    "calhoun_overbids": parse_calhoun_overbids,
}
