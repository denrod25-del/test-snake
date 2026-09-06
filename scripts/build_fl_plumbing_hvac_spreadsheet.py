#!/usr/bin/env python3
"""
Build a Florida plumbing + HVAC company spreadsheet from:
  1) DBPR Construction Industry license extract (official statewide roll)
  2) Optional enrichment from data/companies.json (Google Maps scrape)

Ownership is classified with known PE / public / franchise brand maps plus
name heuristics. Most independents are labeled "Independent / Likely
Family-Owned (unverified)" because Florida does not publish ownership type.

Usage:
  python scripts/build_fl_plumbing_hvac_spreadsheet.py \\
    --dbpr tmp-fl-contractors/CONSTRUCTIONLICENSE_1.csv \\
    --out-dir data/florida-contractors
"""

from __future__ import annotations

import argparse
import csv
import json
import re
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:  # pragma: no cover
    Workbook = None

LICENSE_TYPES = {
    "CAC": ("HVAC", "Certified Air Conditioning Contractor"),
    "RA": ("HVAC", "Registered Air Conditioning Contractor"),
    "CFC": ("Plumbing", "Certified Plumbing Contractor"),
    "RF": ("Plumbing", "Registered Plumbing Contractor"),
    "CMC": ("Mechanical", "Certified Mechanical Contractor"),
    "RM": ("Mechanical", "Registered Mechanical Contractor"),
}

COUNTY_CODES = {
    "11": "Alachua",
    "12": "Baker",
    "13": "Bay",
    "14": "Bradford",
    "15": "Brevard",
    "16": "Broward",
    "17": "Calhoun",
    "18": "Charlotte",
    "19": "Citrus",
    "20": "Clay",
    "21": "Collier",
    "22": "Columbia",
    "23": "Miami-Dade",
    "24": "DeSoto",
    "25": "Dixie",
    "26": "Duval",
    "27": "Escambia",
    "28": "Flagler",
    "29": "Franklin",
    "30": "Gadsden",
    "31": "Gilchrist",
    "32": "Glades",
    "33": "Gulf",
    "34": "Hamilton",
    "35": "Hardee",
    "36": "Hendry",
    "37": "Hernando",
    "38": "Highlands",
    "39": "Hillsborough",
    "40": "Holmes",
    "41": "Indian River",
    "42": "Jackson",
    "43": "Jefferson",
    "44": "Lafayette",
    "45": "Lake",
    "46": "Lee",
    "47": "Leon",
    "48": "Levy",
    "49": "Liberty",
    "50": "Madison",
    "51": "Manatee",
    "52": "Marion",
    "53": "Martin",
    "54": "Monroe",
    "55": "Nassau",
    "56": "Okaloosa",
    "57": "Okeechobee",
    "58": "Orange",
    "59": "Osceola",
    "60": "Palm Beach",
    "61": "Pasco",
    "62": "Pinellas",
    "63": "Polk",
    "64": "Putnam",
    "65": "St. Johns",
    "66": "St. Lucie",
    "67": "Santa Rosa",
    "68": "Sarasota",
    "69": "Seminole",
    "70": "Sumter",
    "71": "Suwannee",
    "72": "Taylor",
    "73": "Union",
    "74": "Volusia",
    "75": "Wakulla",
    "76": "Walton",
    "77": "Washington",
    "78": "Unknown",
    "99": "Unknown",
}

# Confirmed corporate / PE platform brands operating in Florida (substring match
# on normalized company name). Patterns are lowercase.
CORPORATE_BRANDS = [
    # Wrench Group (Leonard Green) — Florida brands
    # Exclude "Get Cool Today" style false positives.
    (r"(?<!get\s)\bcool\s*today\b", "Corporate (PE)", "Wrench Group (Leonard Green & Partners)", "Confirmed"),
    (r"\bplumbing\s*today\b", "Corporate (PE)", "Wrench Group (Leonard Green & Partners)", "Confirmed"),
    (r"\benergy\s*today\b", "Corporate (PE)", "Wrench Group (Leonard Green & Partners)", "Confirmed"),
    (r"\bred\s*cap\s*(plumbing|air|heating|hvac)?\b", "Corporate (PE)", "Wrench Group (Leonard Green & Partners)", "Confirmed"),
    # Florida Cool Inc. was acquired by Wrench and folded into CoolToday (2020).
    (r"^florida\s*cool(\s+inc\.?|\s+llc)?$", "Corporate (PE)", "Wrench Group (Leonard Green & Partners) — Florida Cool / CoolToday", "Confirmed"),
    # Apex Service Partners (Alpine Investors) — Florida brands
    (r"\bbest\s*home\s*services\b", "Corporate (PE)", "Apex Service Partners (Alpine Investors)", "Confirmed"),
    (r"\bfrank\s*gay\b", "Corporate (PE)", "Apex Service Partners (Alpine Investors)", "Confirmed"),
    (r"\bright\s*now\s+(heating|cooling|plumbing)\b", "Corporate (PE)", "Apex Service Partners (Alpine Investors)", "Confirmed"),
    # National PE / public operators
    (r"\bars\s*/\s*rescue\s*rooter\b|\bamerican\s*residential\s*services\b|\brescue\s*rooter\b", "Corporate (PE)", "ARS / Rescue Rooter (GI Partners + Charlesbank)", "Confirmed"),
    (r"(^service\s*experts\b|,\s*a\s+service\s*experts\s+company\b)", "Corporate (PE)", "Service Experts (Brookfield / Enercare)", "Confirmed"),
    (r"\bturnpoint\s*(services|hvac|plumbing)?\b", "Corporate (PE)", "TurnPoint Services (OMERS)", "Confirmed"),
    (r"\bsila\s*services\b|\bsila\s*heating\b", "Corporate (PE)", "Sila Services (Goldman Sachs Alternatives)", "Confirmed"),
    (r"\bcomfort\s*systems\s*usa\b", "Corporate (Public)", "Comfort Systems USA (NYSE: FIX)", "Confirmed"),
    (r"\broto[\s\-]*rooter\b", "Corporate (Public)", "Roto-Rooter / Chemed Corp (NYSE: CHE)", "Confirmed"),
]

FRANCHISE_BRANDS = [
    (r"\bbenjamin\s*franklin\b", "Franchise", "Authority Brands — Benjamin Franklin Plumbing", "Confirmed"),
    (r"\bone\s*hour\b.*\b(heating|air|hvac)\b|\bone\s*hour\s*heating\b", "Franchise", "Authority Brands — One Hour Heating & Air Conditioning", "Confirmed"),
    (r"\bmister\s*sparky\b|\bmr\.?\s*sparky\b", "Franchise", "Authority Brands — Mister Sparky", "Confirmed"),
    (r"\bmr\.?\s*rooter\b", "Franchise", "Neighborly — Mr. Rooter", "Confirmed"),
    (r"\baire\s*serv\b", "Franchise", "Neighborly — Aire Serv", "Confirmed"),
    (r"\bmr\.?\s*electric\b", "Franchise", "Neighborly — Mr. Electric", "Confirmed"),
    (r"\brooter[\s\-]*man\b|\brooterman\b", "Franchise", "Rooter-Man franchise system", "Confirmed"),
    (r"\bdrain\s*doctor\b", "Franchise", "Drain Doctor franchise system", "Pattern match"),
    (r"\bclockwork\b", "Franchise", "Clockwork Home Services franchise family", "Pattern match"),
]

# Explicitly researched family-owned Florida operators (not exhaustive).
FAMILY_KNOWN = [
    (r"\bpeaden\b", "Family-Owned", "Peaden Air Conditioning — family-run (Daws brothers); public about page", "Confirmed"),
    (r"\bacree\b", "Family-Owned", "Acree Plumbing, Air & Electric — long-standing Florida family brand", "Pattern match"),
    (r"\baztec\s*plumbing\b", "Family-Owned", "Aztec Plumbing — long-standing Tampa Bay independent", "Pattern match"),
]

FAMILY_NAME_HINTS = re.compile(
    r"family[\s\-]*owned|family[\s\-]*operated|&\s*sons?\b|and\s+sons?\b|"
    r"\bbrothers\b|\bfather\s*(&|and)\s*son\b|\bson\s*(&|and)\s*father\b",
    re.I,
)

ENTITY_CORP_HINT = re.compile(
    r"\b(inc\.?|incorporated|corp\.?|corporation|llc|l\.l\.c\.|ltd|co\.?)\b",
    re.I,
)


def normalize_name(name: str) -> str:
    s = (name or "").upper()
    s = re.sub(r"[^A-Z0-9\s&]", " ", s)
    s = re.sub(r"\b(LLC|L L C|INC|INCORPORATED|CORP|CORPORATION|CO|LTD|PLLC|PA|PC)\b", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def classify_ownership(name: str) -> tuple[str, str, str]:
    """Return (ownership_type, parent_or_notes, confidence)."""
    n = name or ""
    low = n.lower()
    for pattern, otype, notes, conf in CORPORATE_BRANDS:
        if re.search(pattern, low):
            return otype, notes, conf
    for pattern, otype, notes, conf in FRANCHISE_BRANDS:
        if re.search(pattern, low):
            return otype, notes, conf
    for pattern, otype, notes, conf in FAMILY_KNOWN:
        if re.search(pattern, low):
            return otype, notes, conf
    if FAMILY_NAME_HINTS.search(n):
        return "Family-Owned", "Name/marketing indicates family ownership", "Pattern match"
    return (
        "Independent / Likely Family-Owned (unverified)",
        "No confirmed PE/public/franchise match; most FL contractors are independently owned — ownership not verified",
        "Unverified",
    )


def trade_label(types: set[str]) -> str:
    cats = {LICENSE_TYPES[t][0] for t in types if t in LICENSE_TYPES}
    if cats == {"Plumbing"}:
        return "Plumbing"
    if cats == {"HVAC"}:
        return "HVAC"
    if cats == {"Mechanical"}:
        return "Mechanical"
    if "Plumbing" in cats and ("HVAC" in cats or "Mechanical" in cats):
        return "Plumbing + HVAC"
    if "HVAC" in cats and "Mechanical" in cats:
        return "HVAC / Mechanical"
    return " + ".join(sorted(cats)) or "Unknown"


def load_companies_enrichment(path: Path) -> dict[str, dict]:
    if not path.exists():
        return {}
    data = json.loads(path.read_text(encoding="utf-8"))
    out: dict[str, dict] = {}
    for c in data.get("companies") or []:
        key = normalize_name(c.get("name") or "")
        if not key:
            continue
        # Prefer the location with the most reviews when duplicates exist.
        reviews = ((c.get("platforms") or {}).get("google") or {}).get("reviews") or 0
        prev = out.get(key)
        if prev and (prev.get("_reviews") or 0) >= reviews:
            continue
        website = (c.get("website") or "").split("?")[0].rstrip("/")
        out[key] = {
            "phone": c.get("phone") or "",
            "website": website,
            "google_rating": ((c.get("platforms") or {}).get("google") or {}).get("rating"),
            "google_reviews": reviews,
            "maps_city": c.get("city") or "",
            "maps_county": c.get("county") or "",
            "maps_address": c.get("address") or "",
            "services": ", ".join(c.get("services") or []),
            "_reviews": reviews,
        }
    return out


def parse_dbpr(path: Path) -> list[dict]:
    # Aggregate by normalized business name (or licensee if no business name).
    groups: dict[str, dict] = {}

    with path.open(newline="", encoding="latin-1") as f:
        for row in csv.reader(f):
            if len(row) < 21:
                continue
            lic_type = row[1].strip()
            if lic_type not in LICENSE_TYPES:
                continue
            primary = row[13].strip()
            secondary = row[14].strip()
            if primary != "C" or secondary != "A":
                continue

            licensee = (row[2] or "").strip()
            business = (row[3] or "").strip()
            display = business or licensee or "(unnamed)"
            key = normalize_name(display) or display.upper()

            addr1 = (row[5] or "").strip()
            addr2 = (row[6] or "").strip()
            city = (row[8] or "").strip()
            state = (row[9] or "").strip() or "FL"
            zipc = (row[10] or "").strip()
            county_code = (row[11] or "").strip()
            # Out-of-state / foreign codes
            if county_code.startswith("7") or county_code.startswith("8"):
                county = "Out of State" if county_code.startswith("7") else "Foreign"
            else:
                county = COUNTY_CODES.get(county_code, county_code or "Unknown")
            full_lic = (row[20] or "").strip() or f"{lic_type}{row[12].strip()}"
            exp = (row[17] or "").strip()

            g = groups.get(key)
            if not g:
                g = {
                    "company_name": display,
                    "normalized_name": key,
                    "licensee_names": set(),
                    "license_numbers": set(),
                    "license_types": set(),
                    "license_type_labels": set(),
                    "cities": Counter(),
                    "counties": Counter(),
                    "zips": Counter(),
                    "addresses": Counter(),
                    "expirations": [],
                    "license_count": 0,
                }
                groups[key] = g

            if licensee:
                g["licensee_names"].add(licensee)
            g["license_numbers"].add(full_lic)
            g["license_types"].add(lic_type)
            g["license_type_labels"].add(LICENSE_TYPES[lic_type][1])
            g["license_count"] += 1
            if city:
                g["cities"][city.title() if city.isupper() else city] += 1
            if county:
                g["counties"][county] += 1
            if zipc:
                g["zips"][zipc] += 1
            street = " ".join(x for x in (addr1, addr2) if x)
            if street or city:
                line = ", ".join(
                    x
                    for x in (
                        street.title() if street.isupper() else street,
                        (city.title() if city.isupper() else city),
                        state,
                        zipc,
                    )
                    if x
                )
                g["addresses"][line] += 1
            if exp:
                g["expirations"].append(exp)

    rows = []
    for g in groups.values():
        city = g["cities"].most_common(1)[0][0] if g["cities"] else ""
        county = g["counties"].most_common(1)[0][0] if g["counties"] else ""
        address = g["addresses"].most_common(1)[0][0] if g["addresses"] else ""
        zipc = g["zips"].most_common(1)[0][0] if g["zips"] else ""
        ownership, notes, confidence = classify_ownership(g["company_name"])
        rows.append(
            {
                "company_name": g["company_name"],
                "trade": trade_label(g["license_types"]),
                "ownership_type": ownership,
                "ownership_confidence": confidence,
                "parent_company_or_notes": notes,
                "city": city,
                "county": county,
                "zip": zipc,
                "address": address,
                "license_types": ", ".join(sorted(g["license_types"])),
                "license_type_labels": "; ".join(sorted(g["license_type_labels"])),
                "license_numbers": "; ".join(sorted(g["license_numbers"])),
                "active_license_count": g["license_count"],
                "qualifying_agents_or_licensees": "; ".join(sorted(g["licensee_names"])[:8]),
                "latest_expiration_seen": max(g["expirations"]) if g["expirations"] else "",
                "phone": "",
                "website": "",
                "google_rating": "",
                "google_reviews": "",
                "maps_services": "",
                "enrichment_source": "",
                "data_source": "FL DBPR Construction License Extract (Current + Active)",
            }
        )
    rows.sort(key=lambda r: (r["ownership_type"], r["county"], r["company_name"].upper()))
    return rows


def enrich_rows(rows: list[dict], enrichment: dict[str, dict]) -> int:
    matched = 0
    for r in rows:
        info = enrichment.get(r["normalized_name"] if "normalized_name" in r else normalize_name(r["company_name"]))
        if not info:
            info = enrichment.get(normalize_name(r["company_name"]))
        if not info:
            continue
        matched += 1
        r["phone"] = info.get("phone") or ""
        r["website"] = info.get("website") or ""
        r["google_rating"] = info.get("google_rating") if info.get("google_rating") is not None else ""
        r["google_reviews"] = info.get("google_reviews") if info.get("google_reviews") is not None else ""
        r["maps_services"] = info.get("services") or ""
        r["enrichment_source"] = "data/companies.json (Google Maps scrape)"
        # Fill sparse DBPR address fields from Maps when missing
        if not r["city"] and info.get("maps_city"):
            r["city"] = info["maps_city"]
        if not r["county"] and info.get("maps_county"):
            r["county"] = info["maps_county"]
        if not r["address"] and info.get("maps_address"):
            r["address"] = info["maps_address"]
    return matched


COLUMNS = [
    "company_name",
    "trade",
    "ownership_type",
    "ownership_confidence",
    "parent_company_or_notes",
    "city",
    "county",
    "zip",
    "address",
    "phone",
    "website",
    "google_rating",
    "google_reviews",
    "maps_services",
    "license_types",
    "license_type_labels",
    "license_numbers",
    "active_license_count",
    "qualifying_agents_or_licensees",
    "latest_expiration_seen",
    "enrichment_source",
    "data_source",
]


def write_csv(path: Path, rows: list[dict]) -> None:
    with path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=COLUMNS, extrasaction="ignore")
        w.writeheader()
        for r in rows:
            w.writerow({k: r.get(k, "") for k in COLUMNS})


def style_header(ws) -> None:
    fill = PatternFill("solid", fgColor="1F4E79")
    font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(vertical="center", wrap_text=True)


def autosize(ws, max_width: int = 48) -> None:
    for idx, col in enumerate(ws.columns, start=1):
        length = 0
        for cell in col[:200]:
            val = "" if cell.value is None else str(cell.value)
            length = max(length, min(len(val), max_width))
        ws.column_dimensions[get_column_letter(idx)].width = max(10, length + 2)


def write_xlsx(path: Path, rows: list[dict], summary: dict) -> None:
    if Workbook is None:
        raise RuntimeError("openpyxl is required for Excel export")

    wb = Workbook()

    # Summary
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "Florida Plumbing & HVAC Companies"
    ws["A1"].font = Font(bold=True, size=16)
    ws["A2"] = f"Generated (UTC): {summary['generated']}"
    ws["A3"] = (
        "Primary source: Florida DBPR Construction Industry license extract "
        "(Current + Active licenses only)."
    )
    ws["A4"] = (
        "Ownership note: DBPR does not publish family vs corporate ownership. "
        "Corporate/Franchise labels use confirmed brand maps; everything else is "
        "Independent / Likely Family-Owned (unverified) unless a name heuristic matched."
    )
    ws["A6"] = "Metric"
    ws["B6"] = "Count"
    style_header(ws)
    metrics = [
        ("Total companies (unique business names)", summary["total"]),
        ("Corporate (PE / Public)", summary["corporate"]),
        ("Franchise", summary["franchise"]),
        ("Family-Owned (confirmed or name pattern)", summary["family"]),
        ("Independent / Likely Family-Owned (unverified)", summary["independent"]),
        ("Plumbing-only", summary["plumbing"]),
        ("HVAC-only", summary["hvac"]),
        ("Plumbing + HVAC", summary["both"]),
        ("Mechanical / mixed", summary["mechanical"]),
        ("Enriched with phone/website from companies.json", summary["enriched"]),
    ]
    for i, (k, v) in enumerate(metrics, start=7):
        ws[f"A{i}"] = k
        ws[f"B{i}"] = v
    ws["A18"] = "License types included"
    ws["A19"] = "CAC, RA (air conditioning); CFC, RF (plumbing); CMC, RM (mechanical)"
    ws["A21"] = "This is not a claim of 100% of every Florida plumbing/HVAC business —"
    ws["A22"] = "it is every Current+Active DBPR-licensed contractor in those categories,"
    ws["A23"] = "which is the best official statewide roll available."
    autosize(ws)

    def add_sheet(title: str, subset: list[dict]) -> None:
        sheet = wb.create_sheet(title[:31])
        sheet.append(COLUMNS)
        style_header(sheet)
        for r in subset:
            sheet.append([r.get(c, "") for c in COLUMNS])
        sheet.auto_filter.ref = sheet.dimensions
        sheet.freeze_panes = "A2"
        autosize(sheet)

    add_sheet("All Companies", rows)
    add_sheet(
        "Corporate Owned",
        [r for r in rows if r["ownership_type"].startswith("Corporate")],
    )
    add_sheet("Franchise", [r for r in rows if r["ownership_type"] == "Franchise"])
    add_sheet(
        "Family-Owned",
        [r for r in rows if r["ownership_type"] == "Family-Owned"],
    )
    add_sheet(
        "Independent Likely Family",
        [
            r
            for r in rows
            if r["ownership_type"].startswith("Independent")
        ],
    )

    wb.save(path)


def write_readme(path: Path, summary: dict) -> None:
    path.write_text(
        f"""# Florida Plumbing & HVAC Companies Spreadsheet

Generated: **{summary['generated']}** (UTC)

## What this is

A statewide company list of Florida **plumbing**, **HVAC / air conditioning**, and **mechanical** contractors assembled from the official Florida DBPR Construction Industry license extract (Current + Active only), with optional phone/website enrichment from `data/companies.json`.

| Metric | Count |
| --- | ---: |
| Total unique companies | {summary['total']} |
| Corporate (PE / Public) | {summary['corporate']} |
| Franchise | {summary['franchise']} |
| Family-Owned (confirmed or name pattern) | {summary['family']} |
| Independent / Likely Family-Owned (unverified) | {summary['independent']} |
| Enriched from Google Maps scrape | {summary['enriched']} |

## Files

- `florida-plumbing-hvac-companies.xlsx` — Excel workbook with Summary + filtered sheets
- `florida-plumbing-hvac-companies.csv` — flat CSV of all companies
- `florida-plumbing-hvac-companies-corporate.csv`
- `florida-plumbing-hvac-companies-family.csv` — confirmed/pattern Family-Owned only
- `florida-plumbing-hvac-companies-independent.csv` — unverified independents (majority)

## Ownership honesty

Florida **does not publish** whether a contractor is family-owned or corporate-owned. Labels work like this:

1. **Corporate (PE / Public)** — matched to known roll-up / public brands (Wrench Group, Apex Service Partners, ARS/Rescue Rooter, Service Experts, Roto-Rooter/Chemed, etc.).
2. **Franchise** — matched to national franchise systems (Benjamin Franklin, One Hour, Mr. Rooter, Aire Serv, etc.).
3. **Family-Owned** — confirmed research (e.g. Peaden) or strong name cues (`& Sons`, `Family Owned`, `Brothers`).
4. **Independent / Likely Family-Owned (unverified)** — everyone else. Most Florida contractors fall here; treat as independent until verified.

## How to regenerate

```bash
# Download official extract (needs a normal browser User-Agent)
mkdir -p tmp-fl-contractors
curl -fsSL -A "Mozilla/5.0" \\
  -H "Referer: https://www2.myfloridalicense.com/construction-industry/public-records/" \\
  -o tmp-fl-contractors/CONSTRUCTIONLICENSE_1.csv \\
  "https://www2.myfloridalicense.com/sto/file_download/extracts/CONSTRUCTIONLICENSE_1.csv"

python scripts/build_fl_plumbing_hvac_spreadsheet.py \\
  --dbpr tmp-fl-contractors/CONSTRUCTIONLICENSE_1.csv \\
  --companies data/companies.json \\
  --out-dir data/florida-contractors
```

## License types included

- **CAC / RA** — Certified / Registered Air Conditioning
- **CFC / RF** — Certified / Registered Plumbing
- **CMC / RM** — Certified / Registered Mechanical

Status filter: primary status `C` (Current) + secondary status `A` (Active).
""",
        encoding="utf-8",
    )


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--dbpr", type=Path, required=True)
    ap.add_argument("--companies", type=Path, default=Path("data/companies.json"))
    ap.add_argument("--out-dir", type=Path, default=Path("data/florida-contractors"))
    args = ap.parse_args()

    args.out_dir.mkdir(parents=True, exist_ok=True)

    print("Parsing DBPR extract…")
    rows = parse_dbpr(args.dbpr)
    print(f"  {len(rows)} unique companies")

    print("Loading enrichment…")
    enrichment = load_companies_enrichment(args.companies)
    matched = enrich_rows(rows, enrichment)
    print(f"  enriched {matched} / {len(rows)}")

    # Drop helper key if present
    for r in rows:
        r.pop("normalized_name", None)

    summary = {
        "generated": datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S"),
        "total": len(rows),
        "corporate": sum(1 for r in rows if r["ownership_type"].startswith("Corporate")),
        "franchise": sum(1 for r in rows if r["ownership_type"] == "Franchise"),
        "family": sum(1 for r in rows if r["ownership_type"] == "Family-Owned"),
        "independent": sum(1 for r in rows if r["ownership_type"].startswith("Independent")),
        "plumbing": sum(1 for r in rows if r["trade"] == "Plumbing"),
        "hvac": sum(1 for r in rows if r["trade"] == "HVAC"),
        "both": sum(1 for r in rows if r["trade"] == "Plumbing + HVAC"),
        "mechanical": sum(
            1
            for r in rows
            if r["trade"] not in ("Plumbing", "HVAC", "Plumbing + HVAC")
        ),
        "enriched": matched,
    }

    base = args.out_dir / "florida-plumbing-hvac-companies"
    write_csv(Path(str(base) + ".csv"), rows)
    write_csv(
        Path(str(base) + "-corporate.csv"),
        [r for r in rows if r["ownership_type"].startswith("Corporate")],
    )
    write_csv(
        Path(str(base) + "-franchise.csv"),
        [r for r in rows if r["ownership_type"] == "Franchise"],
    )
    write_csv(
        Path(str(base) + "-family.csv"),
        [r for r in rows if r["ownership_type"] == "Family-Owned"],
    )
    write_csv(
        Path(str(base) + "-independent.csv"),
        [r for r in rows if r["ownership_type"].startswith("Independent")],
    )
    write_xlsx(Path(str(base) + ".xlsx"), rows, summary)
    write_readme(args.out_dir / "README.md", summary)

    print("Wrote:")
    for p in sorted(args.out_dir.iterdir()):
        print(f"  {p} ({p.stat().st_size:,} bytes)")
    print("Summary:", summary)


if __name__ == "__main__":
    main()
