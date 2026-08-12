#!/usr/bin/env python3
"""Rebuild data/signals feeds: recent permits + PBC county-held certificates."""

from __future__ import annotations

import json
import urllib.request
from datetime import date, datetime, timezone
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
PERMIT_DIR = ROOT / "data" / "permits"
OUT_DIR = ROOT / "data" / "signals"
SLUGS = ["west-palm-beach", "boca-raton", "jupiter", "st-lucie-county"]

CERT_PAGE = "https://www.pbctax.gov/taxes/property-tax/tax-certificates-and-deeds/"
# Published XLSX filename changes; prefer the latest link scraped from the page,
# with a known-good fallback.
CERT_FALLBACK_XLSX = (
    "https://www.pbctax.gov/wp-content/uploads/2026/08/County_Certs_PBCTax-8.3.26.xlsx"
)
DELINQUENT_LIST_URL = (
    "https://edition.pagesuite-professional.co.uk/html5/desktop/production/default.aspx"
    "?edid=febca27c-97b1-4d64-ab8a-491b35fb7dfe"
)


def build_recent_permits() -> None:
    rows = []
    for slug in SLUGS:
        path = PERMIT_DIR / f"{slug}.json"
        if not path.exists():
            continue
        data = json.loads(path.read_text())
        by = data.get("permitsByParcel") or {}
        meta_gen = data.get("generated")
        for parcel, permits in by.items():
            if not isinstance(permits, list):
                continue
            for p in permits:
                if not isinstance(p, dict):
                    continue
                date_v = p.get("issuedDate") or p.get("appliedDate") or p.get("finalDate") or ""
                rows.append(
                    {
                        "source": slug,
                        "permitNumber": p.get("permitNumber") or p.get("number"),
                        "parcelId": parcel or p.get("parcelId") or "",
                        "address": p.get("address") or "",
                        "type": p.get("type") or p.get("subtype") or "",
                        "status": p.get("status") or "",
                        "valuation": p.get("valuation"),
                        "date": str(date_v)[:10] if date_v else "",
                        "url": p.get("url") or "",
                        "sourceGenerated": meta_gen,
                    }
                )

    rows.sort(key=lambda r: r["date"] or "", reverse=True)
    top = [r for r in rows if r["date"]][:150] or rows[:150]
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    payload = {
        "generated": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        "note": "Top recent permits from cached Tyler EnerGov scrapes. Verify on municipal portals.",
        "count": len(top),
        "poolSize": len(rows),
        "permits": top,
    }
    (OUT_DIR / "recent-permits.json").write_text(json.dumps(payload, indent=2) + "\n")
    print(f"permits: wrote {len(top)} of {len(rows)}")


def find_cert_xlsx_url() -> str:
    try:
        req = urllib.request.Request(CERT_PAGE, headers={"User-Agent": "DeedScoutBot/1.0"})
        with urllib.request.urlopen(req, timeout=30) as resp:
            html = resp.read().decode("utf-8", errors="ignore")
        import re

        matches = re.findall(
            r"https://www\.pbctax\.gov/wp-content/uploads/[^\s\"']+County_Certs[^\s\"']+\.xlsx",
            html,
            flags=re.I,
        )
        if matches:
            return matches[0]
    except Exception as exc:  # noqa: BLE001
        print("cert page scrape failed:", exc)
    return CERT_FALLBACK_XLSX


def build_county_held_certs() -> None:
    try:
        import openpyxl
    except ImportError as exc:
        raise SystemExit("openpyxl required: pip install openpyxl") from exc

    xlsx_url = find_cert_xlsx_url()
    tmp = Path("/tmp/pbc_county_certs.xlsx")
    req = urllib.request.Request(xlsx_url, headers={"User-Agent": "DeedScoutBot/1.0"})
    with urllib.request.urlopen(req, timeout=60) as resp:
        tmp.write_bytes(resp.read())

    wb = openpyxl.load_workbook(tmp, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    hdr = [str(h).strip() if h else "" for h in rows[0]]
    out = []
    for r in rows[1:]:
        if not r or not r[0]:
            continue
        d = {}
        for i, h in enumerate(hdr):
            v = r[i] if i < len(r) else None
            if isinstance(v, datetime):
                v = v.date().isoformat()
            elif isinstance(v, date):
                v = v.isoformat()
            d[h] = v
        pin = str(d.get("PIN") or "")
        out.append(
            {
                "pin": pin,
                "pinDigits": "".join(c for c in pin if c.isdigit()),
                "saleDate": d.get("Sale Date"),
                "saleYear": d.get("Sale Year"),
                "certificateNumber": str(d.get("Certificate Number") or ""),
                "taxYear": d.get("Tax Year"),
                "billNumber": str(d.get("Bill Number") or ""),
                "faceAmount": d.get("Face Amount"),
                "interestRate": d.get("Interest Rate"),
                "totalAmount": d.get("Total Amt"),
                "legal": d.get("Legal Description") or "",
            }
        )
    out.sort(key=lambda x: float(x.get("faceAmount") or 0), reverse=True)
    payload = {
        "generated": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        "sourceUrl": xlsx_url,
        "sourcePage": CERT_PAGE,
        "delinquentListUrl": DELINQUENT_LIST_URL,
        "note": (
            "Palm Beach County held tax certificates published by the Constitutional Tax Collector. "
            "Confirm status with the Tax Collector before purchase. Not a complete county delinquency roll."
        ),
        "count": len(out),
        "certificates": out,
    }
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    (OUT_DIR / "pbc-county-held-certs.json").write_text(json.dumps(payload, indent=2) + "\n")
    print(f"certs: wrote {len(out)} from {xlsx_url}")


def main() -> None:
    build_recent_permits()
    build_county_held_certs()


if __name__ == "__main__":
    main()
